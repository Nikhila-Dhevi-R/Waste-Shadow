from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, request, send_from_directory
from mongita import MongitaClientDisk
import numpy as np
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader
from rapidocr_onnxruntime import RapidOCR
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
DB_DIR = BASE_DIR / ".mongita"
UPLOAD_DIR.mkdir(exist_ok=True)
DB_DIR.mkdir(exist_ok=True)

app = Flask(__name__, static_folder=str(BASE_DIR), static_url_path="")

client = MongitaClientDisk(host=str(DB_DIR))
db = client["wasteshadow"]
users_collection = db["users"]
purchases_collection = db["purchases"]
waste_profiles_collection = db["wasteProfiles"]
predictions_collection = db["predictions"]
ocr_engine = RapidOCR()


PRODUCT_PROFILES = [
    {
        "category": "Cleaning",
        "packaging": "Multi-layer plastic",
        "keywords": ["detergent", "cleaner", "dishwash", "soap", "bleach", "sanitizer", "phenyl"],
        "visible_per_unit": 0.15,
        "invisible_per_unit": 1.28,
        "water_per_unit": 78,
        "recommendation": "Choose refill packs or concentrated cleaners to reduce packaging layers and transport waste.",
    },
    {
        "category": "Beverages",
        "packaging": "Plastic bottle",
        "keywords": ["juice", "soda", "cola", "drink", "water bottle", "tea", "coffee", "milkshake"],
        "visible_per_unit": 0.08,
        "invisible_per_unit": 0.88,
        "water_per_unit": 54,
        "recommendation": "Prefer larger reusable containers or locally bottled beverages to cut repeated packaging waste.",
    },
    {
        "category": "Snacks",
        "packaging": "Single-use plastic",
        "keywords": ["chips", "biscuit", "cookie", "cookies", "cracker", "crackers", "snack", "namkeen", "chocolate", "candy"],
        "visible_per_unit": 0.04,
        "invisible_per_unit": 0.73,
        "water_per_unit": 26,
        "recommendation": "Swap individually packed snacks for bulk options to lower the invisible waste index quickly.",
    },
    {
        "category": "Dairy",
        "packaging": "Composite carton",
        "keywords": ["milk", "curd", "paneer", "cheese", "butter", "yogurt", "ghee"],
        "visible_per_unit": 0.07,
        "invisible_per_unit": 0.96,
        "water_per_unit": 112,
        "recommendation": "Buy dairy in returnable or bulk packaging where available to reduce carton and cold-chain impact.",
    },
    {
        "category": "Produce",
        "packaging": "Minimal packaging",
        "keywords": ["apple", "banana", "onion", "tomato", "potato", "carrot", "vegetable", "fruit", "spinach"],
        "visible_per_unit": 0.01,
        "invisible_per_unit": 0.22,
        "water_per_unit": 38,
        "recommendation": "Seasonal local produce keeps transport materials and storage losses lower than imported items.",
    },
    {
        "category": "Grains",
        "packaging": "Poly sack",
        "keywords": ["rice", "wheat", "flour", "atta", "dal", "lentil", "oats", "cereal"],
        "visible_per_unit": 0.05,
        "invisible_per_unit": 0.47,
        "water_per_unit": 68,
        "recommendation": "Choose refill counters or larger staple packs to spread packaging impact over more servings.",
    },
    {
        "category": "Personal Care",
        "packaging": "Plastic tube",
        "keywords": ["shampoo", "toothpaste", "lotion", "cream", "conditioner", "face wash", "deodorant"],
        "visible_per_unit": 0.09,
        "invisible_per_unit": 1.05,
        "water_per_unit": 41,
        "recommendation": "Refill stations and solid alternatives can significantly lower personal care packaging waste.",
    },
    {
        "category": "Disposable Care",
        "packaging": "Single-use plastic",
        "keywords": ["baby wipes", "wipes", "wet wipes", "diaper", "diapers", "sanitary pads", "pads", "liners"],
        "visible_per_unit": 0.11,
        "invisible_per_unit": 1.18,
        "water_per_unit": 36,
        "recommendation": "Choose reusable cloth wipes, refill packs, or plastic-light alternatives to reduce single-use disposal waste.",
    },
    {
        "category": "Household",
        "packaging": "Mixed packaging",
        "keywords": ["tissue", "foil", "paper towel", "bag", "wrap", "container", "garbage", "plastic", "packet", "pouch", "disposable"],
        "visible_per_unit": 0.12,
        "invisible_per_unit": 0.83,
        "water_per_unit": 24,
        "recommendation": "Reusable wraps and durable household substitutes reduce repeated mixed-packaging waste.",
    },
]

DEFAULT_PROFILE = {
    "category": "General",
    "packaging": "Mixed packaging",
    "visible_per_unit": 0.06,
    "invisible_per_unit": 0.61,
    "water_per_unit": 33,
    "recommendation": "Review this product for reusable, refillable, or low-packaging alternatives.",
}

DEMO_ITEMS = [
    {"product": "Detergent 1kg", "quantity": 1},
    {"product": "Potato chips", "quantity": 3},
    {"product": "Milk 1L", "quantity": 2},
    {"product": "Rice bag 5kg", "quantity": 1},
    {"product": "Shampoo bottle", "quantity": 1},
    {"product": "Tomato", "quantity": 4},
]


def ensure_default_user() -> None:
    if not users_collection.find_one({"userName": "Guest User"}):
        users_collection.insert_one(
            {
                "userName": "Guest User",
                "createdAt": datetime.utcnow().isoformat(),
                "preferredFocus": "Preventive sustainability",
            }
        )


def normalize_purchase_date(raw_date: str | None) -> str:
    if not raw_date:
        return datetime.now().date().isoformat()

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw_date, fmt).date().isoformat()
        except ValueError:
            continue

    return datetime.now().date().isoformat()


def parse_quantity(raw_value: Any) -> float:
    if raw_value is None:
        return 1.0

    if isinstance(raw_value, (int, float)):
        return max(float(raw_value), 1.0)

    text = str(raw_value).strip().lower()
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    return max(float(match.group(1)), 1.0) if match else 1.0


def choose_profile(product_name: str) -> dict[str, Any]:
    lowered = product_name.lower()
    for profile in PRODUCT_PROFILES:
        if any(re.search(rf"\b{re.escape(keyword.lower())}\b", lowered) for keyword in profile["keywords"]):
            return profile
    return DEFAULT_PROFILE


def normalize_item(product_name: str, quantity: Any) -> dict[str, Any]:
    qty = parse_quantity(quantity)
    profile = choose_profile(product_name)

    visible = round(profile["visible_per_unit"] * qty, 3)
    invisible = round(profile["invisible_per_unit"] * qty, 3)
    water = round(profile["water_per_unit"] * qty, 2)
    index = round((invisible * 10) + (visible * 4) + (water / 15), 2)

    return {
        "product": product_name,
        "quantity": qty,
        "category": profile["category"],
        "packagingType": profile["packaging"],
        "visibleWasteKg": visible,
        "invisibleWasteKg": invisible,
        "waterFootprintL": water,
        "invisibleWasteIndex": index,
        "recommendation": profile["recommendation"],
    }


def parse_csv_file(file_storage) -> list[dict[str, Any]]:
    text = file_storage.read().decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    items = []

    for row in reader:
        product = row.get("product") or row.get("item") or row.get("name")
        if not product:
            continue
        quantity = row.get("quantity") or row.get("qty") or row.get("count") or 1
        items.append({"product": product.strip(), "quantity": quantity})

    return items


def parse_xlsx_file(file_storage) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_storage.read()), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    items = []
    for row in rows[1:]:
        row_dict = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
        product = row_dict.get("product") or row_dict.get("item") or row_dict.get("name")
        if not product:
            continue
        quantity = row_dict.get("quantity") or row_dict.get("qty") or row_dict.get("count") or 1
        items.append({"product": str(product).strip(), "quantity": quantity})

    return items


def parse_pdf_file(file_storage) -> list[dict[str, Any]]:
    reader = PdfReader(io.BytesIO(file_storage.read()))
    items = []
    ignore_tokens = {
        "invoice", "bill", "total", "amount", "subtotal", "gst", "tax", "discount", "date",
        "qty", "quantity", "price", "rate", "cash", "upi", "phone", "mobile", "thank you"
    }

    for page in reader.pages:
        text = page.extract_text() or ""
        items.extend(extract_items_from_text_lines(text.splitlines(), ignore_tokens))

    return items


def extract_items_from_text_lines(lines: list[str], ignore_tokens: set[str] | None = None) -> list[dict[str, Any]]:
    items = []
    ignore_tokens = {
        "invoice", "bill", "total", "amount", "subtotal", "gst", "tax", "discount", "date",
        "qty", "quantity", "price", "rate", "cash", "upi", "phone", "mobile", "thank you"
    }
    if ignore_tokens:
        ignore_tokens = ignore_tokens | set(ignore_tokens)

    for line in lines:
        cleaned = re.sub(r"\s+", " ", line).strip()
        if not cleaned or len(cleaned) < 3:
            continue

        lowered = cleaned.lower()
        if any(token in lowered for token in ignore_tokens):
            continue

        patterns = [
            r"^(?P<product>[A-Za-z][A-Za-z0-9\s&\-/(),.%]+?)\s+x\s*(?P<qty>\d+(?:\.\d+)?)\b",
            r"^(?P<product>[A-Za-z][A-Za-z0-9\s&\-/(),.%]+?)\s{2,}(?P<qty>\d+(?:\.\d+)?)\b",
            r"^(?P<product>[A-Za-z][A-Za-z0-9\s&\-/(),.%]+?)\s+(?P<qty>\d+(?:\.\d+)?)\s+\d",
            r"^(?P<product>[A-Za-z][A-Za-z0-9\s&\-/(),.%]+?)\s+(?P<qty>\d+(?:\.\d+)?)$",
        ]

        match = None
        for pattern in patterns:
            match = re.match(pattern, cleaned)
            if match:
                product = re.sub(r"[\.\-_]{2,}", " ", match.group("product")).strip(" -.")
                if len(product) >= 3:
                    items.append({"product": product, "quantity": match.group("qty")})
                break

        if match:
            continue

        token_match = re.match(
            r"^(?P<product>[A-Za-z][A-Za-z0-9\s&\-/(),.%]+?)\s+(?P<qty>\d+(?:\.\d+)?)\s+(?:rs\.?|inr|\d[\d.,]*)",
            cleaned,
            flags=re.IGNORECASE,
        )
        if token_match:
            product = token_match.group("product").strip(" -.")
            if len(product) >= 3:
                items.append({"product": product, "quantity": token_match.group("qty")})

    return items


def parse_image_file(file_storage) -> list[dict[str, Any]]:
    image = Image.open(io.BytesIO(file_storage.read())).convert("RGB")
    image_array = np.array(image)
    ocr_result, _ = ocr_engine(image_array)
    if not ocr_result:
        return []

    rows = []
    ignore_tokens = {
        "invoice", "bill", "total", "amount", "subtotal", "gst", "tax", "discount", "date",
        "qty", "quantity", "price", "rate", "cash", "upi", "phone", "mobile", "thank you",
        "store", "shop", "chicago", "illinois"
    }

    for entry in ocr_result:
        if len(entry) < 2 or not entry[1]:
            continue
        box, text = entry[0], str(entry[1]).strip()
        x_values = [point[0] for point in box]
        y_values = [point[1] for point in box]
        rows.append(
            {
                "text": text,
                "x": sum(x_values) / len(x_values),
                "y": sum(y_values) / len(y_values),
            }
        )

    extracted_items = []
    used_indexes: set[int] = set()
    price_like = re.compile(r"^\$?\d+(?:\.\d{2})?$")

    for index, row in enumerate(rows):
        lowered = row["text"].lower()
        if any(token in lowered for token in ignore_tokens):
            continue

        if price_like.match(row["text"]):
            continue

        nearest_price_index = None
        nearest_price_distance = None
        for candidate_index, candidate in enumerate(rows):
            if candidate_index == index or candidate_index in used_indexes:
                continue
            if not price_like.match(candidate["text"]):
                continue

            vertical_gap = abs(candidate["y"] - row["y"])
            horizontal_gap = candidate["x"] - row["x"]
            if vertical_gap <= 18 and horizontal_gap > 120:
                score = vertical_gap + (horizontal_gap / 1000)
                if nearest_price_distance is None or score < nearest_price_distance:
                    nearest_price_index = candidate_index
                    nearest_price_distance = score

        if nearest_price_index is not None:
            product = row["text"].replace("1lb", "1 lb").replace("11b", "1 lb")
            extracted_items.append({"product": product, "quantity": 1})
            used_indexes.add(index)
            used_indexes.add(nearest_price_index)

    if extracted_items:
        return extracted_items

    lines = [entry["text"] for entry in rows]
    return extract_items_from_text_lines(lines, ignore_tokens)


def load_items_from_upload(file_storage) -> list[dict[str, Any]]:
    suffix = Path(file_storage.filename).suffix.lower()

    if suffix == ".csv":
        return parse_csv_file(file_storage)
    if suffix == ".xlsx":
        return parse_xlsx_file(file_storage)
    if suffix == ".pdf":
        return parse_pdf_file(file_storage)
    if suffix in {".png", ".jpg", ".jpeg"}:
        return parse_image_file(file_storage)

    raise ValueError("Unsupported file type. Use CSV, XLSX, PDF, JPG, JPEG, or PNG.")


def analyze_purchase_items(items: list[dict[str, Any]]) -> dict[str, Any]:
    analyzed_items = [normalize_item(item["product"], item.get("quantity")) for item in items if item.get("product")]
    if not analyzed_items:
        raise ValueError("No valid product rows were found in the uploaded file.")

    total_visible = round(sum(item["visibleWasteKg"] for item in analyzed_items), 3)
    total_invisible = round(sum(item["invisibleWasteKg"] for item in analyzed_items), 3)
    total_water = round(sum(item["waterFootprintL"] for item in analyzed_items), 2)
    total_index = round(sum(item["invisibleWasteIndex"] for item in analyzed_items) / len(analyzed_items), 2)

    category_totals: dict[str, dict[str, float]] = defaultdict(lambda: {"visible": 0.0, "invisible": 0.0})
    for item in analyzed_items:
        category_totals[item["category"]]["visible"] += item["visibleWasteKg"]
        category_totals[item["category"]]["invisible"] += item["invisibleWasteKg"]

    category_breakdown = sorted(
        [
            {
                "category": category,
                "visibleWasteKg": round(values["visible"], 3),
                "invisibleWasteKg": round(values["invisible"], 3),
            }
            for category, values in category_totals.items()
        ],
        key=lambda entry: entry["invisibleWasteKg"],
        reverse=True,
    )

    recommendations = []
    seen_titles: set[str] = set()

    plastic_priority_items = [
        item
        for item in analyzed_items
        if "plastic" in item["packagingType"].lower()
    ]
    ranked_items = sorted(analyzed_items, key=lambda entry: entry["invisibleWasteKg"], reverse=True)

    for item in plastic_priority_items + ranked_items:
        title = f"{item['product']} alternative"
        if title in seen_titles:
            continue

        recommendations.append(
            {
                "title": title,
                "message": f"{item['product']} uses {item['packagingType'].lower()}. {item['recommendation']}",
            }
        )
        seen_titles.add(title)

        if len(recommendations) >= 5:
            break

    generated_at = datetime.now().strftime("%d %b %Y, %I:%M %p")
    return {
        "items": analyzed_items,
        "summary": {
            "totalVisibleWasteKg": total_visible,
            "totalInvisibleWasteKg": total_invisible,
            "totalWaterFootprintL": total_water,
            "invisibleWasteIndex": total_index,
            "generatedAt": generated_at,
        },
        "categoryBreakdown": category_breakdown,
        "recommendations": recommendations,
    }


def build_predictions() -> list[dict[str, Any]]:
    purchases = list(purchases_collection.find())
    if not purchases:
        return [
            {"label": "30 Days", "days": 30, "projectedInvisibleWasteKg": 0.0},
            {"label": "90 Days", "days": 90, "projectedInvisibleWasteKg": 0.0},
            {"label": "365 Days", "days": 365, "projectedInvisibleWasteKg": 0.0},
        ]

    sorted_purchases = sorted(purchases, key=lambda item: item["purchaseDate"])
    first_date = datetime.fromisoformat(sorted_purchases[0]["purchaseDate"])
    last_date = datetime.fromisoformat(sorted_purchases[-1]["purchaseDate"])
    span_days = max((last_date - first_date).days + 1, 7)

    total_hidden = sum(entry["summary"]["totalInvisibleWasteKg"] for entry in sorted_purchases)
    average_daily_hidden = total_hidden / span_days
    trend_factor = 1 + min(len(sorted_purchases) - 1, 6) * 0.04

    predictions = []
    for days in (30, 90, 365):
        projected = round(average_daily_hidden * days * trend_factor, 2)
        predictions.append({"label": f"{days} Days", "days": days, "projectedInvisibleWasteKg": projected})

    return predictions


def serialize_purchase(purchase: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": str(purchase.get("_id", "")),
        "fileName": purchase["fileName"],
        "userName": purchase["userName"],
        "uploadedAt": purchase["uploadedAtDisplay"],
        "totalInvisibleWasteKg": purchase["summary"]["totalInvisibleWasteKg"],
        "totalWaterFootprintL": purchase["summary"]["totalWaterFootprintL"],
    }


def refresh_waste_profile(user_name: str) -> None:
    purchases = list(purchases_collection.find({"userName": user_name}))
    if not purchases:
        return

    total_visible = sum(item["summary"]["totalVisibleWasteKg"] for item in purchases)
    total_invisible = sum(item["summary"]["totalInvisibleWasteKg"] for item in purchases)
    total_water = sum(item["summary"]["totalWaterFootprintL"] for item in purchases)

    waste_profiles_collection.delete_many({"userName": user_name})
    waste_profiles_collection.insert_one(
        {
            "userName": user_name,
            "analyses": len(purchases),
            "cumulativeVisibleWasteKg": round(total_visible, 3),
            "cumulativeInvisibleWasteKg": round(total_invisible, 3),
            "cumulativeWaterFootprintL": round(total_water, 2),
            "updatedAt": datetime.utcnow().isoformat(),
        }
    )


def persist_analysis(
    analysis: dict[str, Any],
    file_name: str,
    file_type: str,
    user_name: str,
    purchase_date: str,
) -> dict[str, Any]:
    analysis["summary"]["fileName"] = file_name

    purchase_doc = {
        "fileName": file_name,
        "fileType": file_type,
        "userName": user_name,
        "purchaseDate": purchase_date,
        "uploadedAt": datetime.utcnow().isoformat(),
        "uploadedAtDisplay": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "items": analysis["items"],
        "summary": analysis["summary"],
        "categoryBreakdown": analysis["categoryBreakdown"],
        "recommendations": analysis["recommendations"],
    }
    purchases_collection.insert_one(purchase_doc)
    refresh_waste_profile(user_name)

    latest_predictions = build_predictions()
    predictions_collection.insert_one(
        {
            "userName": user_name,
            "createdAt": datetime.utcnow().isoformat(),
            "predictions": latest_predictions,
        }
    )

    analysis["predictions"] = latest_predictions
    return analysis


def clear_all_history() -> None:
    purchases_collection.delete_many({})
    predictions_collection.delete_many({})
    waste_profiles_collection.delete_many({})


ensure_default_user()


@app.get("/")
def root() -> Any:
    return send_from_directory(BASE_DIR, "index.html")


@app.get("/api/overview")
def overview() -> Any:
    purchases = sorted(list(purchases_collection.find()), key=lambda item: item["uploadedAt"], reverse=True)
    total_analyses = len(purchases)
    total_invisible = round(sum(item["summary"]["totalInvisibleWasteKg"] for item in purchases), 3)
    total_water = round(sum(item["summary"]["totalWaterFootprintL"] for item in purchases), 2)

    latest = None
    if purchases:
        latest_purchase = purchases[0]
        prediction_doc = predictions_collection.find_one(sort=[("createdAt", -1)])
        latest = {
            "summary": latest_purchase["summary"],
            "items": latest_purchase["items"],
            "categoryBreakdown": latest_purchase["categoryBreakdown"],
            "recommendations": latest_purchase["recommendations"],
            "predictions": prediction_doc["predictions"] if prediction_doc else build_predictions(),
        }

    return jsonify(
        {
            "totalAnalyses": total_analyses,
            "cumulativeInvisibleWasteKg": total_invisible,
            "cumulativeWaterFootprintL": total_water,
            "recentAnalyses": [serialize_purchase(item) for item in purchases[:5]],
            "latestAnalysis": latest,
        }
    )


@app.post("/api/demo")
def demo() -> Any:
    analysis = analyze_purchase_items(DEMO_ITEMS)
    response = persist_analysis(
        analysis=analysis,
        file_name="demo_analysis.csv",
        file_type="csv",
        user_name="Guest User",
        purchase_date=datetime.now().date().isoformat(),
    )
    return jsonify(response)


@app.delete("/api/history")
def clear_history() -> Any:
    clear_all_history()
    return jsonify({"status": "ok"})


@app.post("/api/analyze")
def analyze() -> Any:
    upload = request.files.get("file")
    user_name = (request.form.get("user_name") or "Guest User").strip() or "Guest User"
    purchase_date = normalize_purchase_date(request.form.get("purchase_date"))

    if upload is None or not upload.filename:
        return jsonify({"error": "Please choose a file to analyze."}), 400

    filename = secure_filename(upload.filename)
    suffix = Path(filename).suffix.lower().lstrip(".")

    try:
        raw_bytes = upload.read()
        saved_path = UPLOAD_DIR / filename
        saved_path.write_bytes(raw_bytes)
        upload.stream.seek(0)

        items = load_items_from_upload(upload)
        analysis = analyze_purchase_items(items)
        response = persist_analysis(
            analysis=analysis,
            file_name=filename,
            file_type=suffix,
            user_name=user_name,
            purchase_date=purchase_date,
        )
        return jsonify(response)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception as error:
        return jsonify({"error": f"Analysis failed: {error}"}), 500


if __name__ == "__main__":
    ensure_default_user()
    app.run(debug=True)
